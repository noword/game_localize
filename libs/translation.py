from pathlib import Path
import re
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl import worksheet
from dataclasses import dataclass
from collections import OrderedDict
import json


@dataclass
class Theme:
    font_name: str

    header_background_color_start: str
    header_background_color_end: str
    header_color: str

    original_background_color_start: str
    original_background_color_end: str
    original_color: str

    translation_background_color_start: str
    translation_background_color_end: str
    translation_color: str

    comment_background_color_start: str
    comment_background_color_end: str
    comment_color: str

    other_background_color_start: str
    other_background_color_end: str
    other_color: str

    def __post_init__(self):
        self.header_fill = openpyxl.styles.GradientFill(stop=(self.header_background_color_start,
                                                              self.header_background_color_end))
        self.header_font = openpyxl.styles.Font(name=self.font_name, bold=True, color=self.header_color)

        self.org_fill = openpyxl.styles.GradientFill(stop=(self.original_background_color_start,
                                                           self.original_background_color_end))
        self.org_font = openpyxl.styles.Font(name=self.font_name, color=self.original_color)

        self.trans_fill = openpyxl.styles.GradientFill(stop=(self.translation_background_color_start,
                                                             self.translation_background_color_end))
        self.trans_font = openpyxl.styles.Font(name=self.font_name, color=self.translation_color)

        self.comment_fill = openpyxl.styles.GradientFill(stop=(self.comment_background_color_start,
                                                               self.comment_background_color_end))
        self.comment_font = openpyxl.styles.Font(name=self.font_name, color=self.comment_color)

        self.other_fill = openpyxl.styles.GradientFill(stop=(self.other_background_color_start,
                                                             self.other_background_color_end))
        self.other_font = openpyxl.styles.Font(name=self.font_name, color=self.other_color)

        self.font = openpyxl.styles.Font(name=self.font_name)
        self.border = Border(left=Side(style='hair'),
                             right=Side(style='hair'),
                             top=Side(style='hair'),
                             bottom=Side(style='hair'),
                             )


THEME = Theme(font_name='Consolas',
              header_background_color_start='FCF3CF',
              header_background_color_end='FEF9E7',
              header_color='333300',
              original_background_color_start='D6EAF8',
              original_background_color_end='EAFAF1',
              original_color='000033',
              translation_background_color_start='D5F5E3',
              translation_background_color_end='EAFAF1',
              translation_color='003300',
              comment_background_color_start='FADBD8',
              comment_background_color_end='FDEDEC',
              comment_color='330000',
              other_background_color_start='FFFFFF',
              other_background_color_end='F8F8F8',
              other_color='330000'
              )


class Translation(list):
    def __init__(self,
                 name: str = None,
                 theme: Theme = THEME):
        self.theme = theme
        if name is not None:
            self.load(name)

    def load(self, name: str):
        LOAD = {'.xlsx': self.__load_xlsx,
                # '.csv': self.__load_csv,
                '.json': self.__load_json
                }
        suffix = Path(name).suffix
        if suffix in LOAD:
            LOAD[suffix](name)
        else:
            raise TypeError(f'unsupported file type: {suffix}')

    def save(self, name: str, index: str):
        SAVE = {'.xlsx': self.__save_xlsx,
                # '.csv': self.__save_csv,
                '.json': self.__save_json
                }

        self.__before_save()
        suffix = Path(name).suffix
        if suffix in SAVE:
            SAVE[suffix](name, index)
        else:
            raise TypeError(f'unsupported file type: {suffix}')

    def __before_save(self):
        keys = list(self[0].keys())
        with_no = 'No.' in keys
        with_comment = 'Comments' in keys
        for i, d in enumerate(self, start=1):
            d = d.copy()
            row = self[i - 1] = {}
            if not with_no:
                row['No.'] = i
            row.update(d)
            if not with_comment:
                row['Comments'] = ''

    def __load_xlsx(self, name: str):
        wb = openpyxl.load_workbook(name)
        ws = wb.active
        keys = [r.value for r in next(ws.rows)]
        for row in ws.iter_rows(min_row=2):
            d = OrderedDict(zip(keys, [r.value for r in row]))
            for key, value in d.items():
                if isinstance(value, float) or isinstance(value, int):
                    d[key] = str(value)
                elif d[key] is None:
                    d[key] = ""
            self.append(d)

    def __save_xlsx(self, name: str, index: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        keys = list(self[0].keys())
        ws.append(keys)

        for d in self:
            ws.append(list(d.values()))

        self.__set_excel_styles(ws, keys.index(index) + 1 if index in keys else 1)
        ws.freeze_panes = ws['A2']
        wb.save(name)

    def __load_json(self, name):
        for t in json.load(open(name, encoding='utf-8')):
            self.append(t)

    def __save_json(self, name: str, index: str):
        json.dump(self, open(name, 'w'), encoding='utf-8', ensure_ascii=False, indent=4)

    def __set_excel_styles(self, ws: worksheet, frezze_index: int):
        trans_rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual',
                                                         formula=['""'],
                                                         border=self.theme.border,
                                                         fill=self.theme.trans_fill,
                                                         font=self.theme.trans_font)
        comment_rule = openpyxl.formatting.rule.CellIsRule(operator='notEqual',
                                                           formula=['""'],
                                                           border=self.theme.border,
                                                           fill=self.theme.comment_fill,
                                                           font=self.theme.comment_font)

        for i, row in enumerate(ws.iter_rows()):
            for cell in row:
                if cell.value is None:
                    cell.value = ''
                elif not isinstance(cell.value, str):
                    cell.value = str(cell.value)
                cell.number_format = '@'
                cell.data_type = 's'
                cell.quotePrefix = True

                if i == 0:
                    cell.font = self.theme.header_font
                    cell.fill = self.theme.header_fill
                else:
                    cell.font = self.theme.font

        # the comments style
        comment_letter = openpyxl.utils.get_column_letter(ws.max_column)
        if ws[comment_letter + '1'].value == 'Comments':
            ws.conditional_formatting.add(f'{comment_letter}2:{comment_letter}{ws.max_row}', comment_rule)
            cond_end = openpyxl.utils.get_column_letter(ws.max_column - 1)
        else:
            cond_end = openpyxl.utils.get_column_letter(ws.max_column)

        # the oterh style
        end = openpyxl.utils.get_column_letter(frezze_index - 1)
        index_cells = ws[f'A2:{end}{ws.max_row}']
        for row in index_cells:
            for cell in row:
                cell.fill = self.theme.other_fill
                cell.font = self.theme.other_font
                cell.border = self.theme.border

        # the translation style
        cond_start = openpyxl.utils.get_column_letter(frezze_index + 1)
        ws.conditional_formatting.add(f'{cond_start}2:{cond_end}{ws.max_row}', trans_rule)

        # the original style
        start = openpyxl.utils.get_column_letter(frezze_index)
        index_cells = ws[f'{start}2:{start}{ws.max_row}']
        for row in index_cells:
            row[0].fill = self.theme.org_fill
            row[0].font = self.theme.org_font
            row[0].border = self.theme.border

    def get_trans(self, index: str):
        return dict([(d[index], d) for d in self])

    def check_size(self, org: str, trans: str, encoding='utf-8'):
        result = []
        for i, d in enumerate(self):
            if len(d[org].encode(encoding)) < len(d[trans].encode(encoding)):
                result.append(d)
        return result

    def check_vars(self, regex: str, org: str, trans: str, ordered: bool = True):
        resuls = []
        for i, d in enumerate(self, start=1):
            org_vars = re.findall(regex, d[org])
            trans_vars = re.findall(regex, d[trans])
            if len(org_vars) != len(trans_vars):
                resuls.append([i, org_vars, trans_vars])
            else:
                if not ordered:
                    unordered_org_vars = sorted(org_vars)
                    unordered_trans_vars = sorted(trans_vars)
                    if unordered_org_vars != unordered_trans_vars:
                        resuls.append([i, org_vars, trans_vars])
                elif org_vars != trans_vars:
                    resuls.append([i, org_vars, trans_vars])
        return resuls


if __name__ == '__main__':
    trans = Translation()
    trans.append({'en': 'hello world', 'cn': '测试 xxx123xxx'})
    trans.save('example.xlsx', index='en')
